#!/usr/bin/env python3
"""Export 16-quarter Business Trend Dashboard metrics for the AI beneficiary list."""

from __future__ import annotations

import argparse
import math
import re
import sys
from pathlib import Path
from typing import Any, Optional

PROJECT_ROOT = Path(__file__).resolve().parents[1]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

import numpy as np
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo

from core_backend import (
    BUSINESS_QUARTER_TREND_DEFAULT_WEIGHTS,
    calculate_business_quarter_trend_details,
    read_df,
)
from db_session import DbCompat, SessionLocal


QUARTER_RANGE = 16

SHEET1_COLUMNS = [
    "Company",
    "Ticker",
    "Business Quarter Trend Score",
    "Weighted Median Revenue Growth",
    "Weighted Median Operating Margin",
    "Weighted Median Operating Margin Change",
    "Weighted Median Incremental Operating Margin",
    "Weighted Median Days Sales Outstanding",
    "Weighted Median Capex / OCF",
]

METRIC_COLUMNS = SHEET1_COLUMNS[2:]
PERCENT_COLUMNS = {
    "Weighted Median Revenue Growth",
    "Weighted Median Operating Margin",
    "Weighted Median Operating Margin Change",
    "Weighted Median Incremental Operating Margin",
    "Weighted Median Capex / OCF",
}

QUARTERLY_TABLES = {
    "revenue": ("revenues_quarterly", "revenue"),
    "operating_income": ("operating_income_quarterly", "operating_income"),
    "deferred_revenue": ("deferred_revenue_quarterly", "deferred_revenue"),
    "accounts_receivable": ("accounts_receivable_quarterly", "accounts_receivable"),
    "capital_expenditures": ("capital_expenditures_quarterly", "capital_expenditures"),
    "operating_cash_flow": ("operating_cash_flow_quarterly", "operating_cash_flow"),
}


def normalized_ticker(value: Any) -> str:
    return str(value or "").strip().upper().replace(".", "-")


def normalized_name(value: Any) -> str:
    return re.sub(r"[^a-z0-9]+", "", str(value or "").lower())


def finite_or_none(value: Any) -> Optional[float]:
    if value is None:
        return None
    try:
        number = float(value)
    except (TypeError, ValueError):
        return None
    return number if math.isfinite(number) else None


def load_source_companies(source_path: Path) -> list[dict[str, Any]]:
    workbook = load_workbook(source_path, read_only=True, data_only=True)
    worksheet = workbook["Companies"]
    rows: list[dict[str, Any]] = []
    for values in worksheet.iter_rows(min_row=6, values_only=True):
        rank, company, ticker = values[:3]
        if rank is None:
            continue
        rows.append({"rank": int(rank), "company": str(company).strip(), "ticker": str(ticker).strip()})
    workbook.close()
    return rows


def match_company_ids(source_rows: list[dict[str, Any]], companies: pd.DataFrame) -> None:
    companies = companies.copy()
    companies["ticker_norm"] = companies["ticker"].map(normalized_ticker)
    companies["name_norm"] = companies["name"].map(normalized_name)
    for source in source_rows:
        candidates = companies[companies["ticker_norm"] == normalized_ticker(source["ticker"])]
        if len(candidates) > 1:
            exact_name = candidates[candidates["name_norm"] == normalized_name(source["company"])]
            if len(exact_name) == 1:
                candidates = exact_name
        source["company_id"] = int(candidates.iloc[0]["id"]) if len(candidates) == 1 else None


def load_quarterly_inputs(
    conn: DbCompat,
    company_ids: list[int],
) -> dict[str, dict[int, dict[str, float]]]:
    placeholders = ",".join(["?"] * len(company_ids))
    loaded: dict[str, dict[int, dict[str, float]]] = {}
    for input_key, (table, value_col) in QUARTERLY_TABLES.items():
        frame = read_df(
            f"""
            SELECT company_id, quarter_end, {value_col}
            FROM {table}
            WHERE company_id IN ({placeholders})
            ORDER BY company_id, quarter_end DESC
            """,
            conn,
            params=company_ids,
        )
        by_company: dict[int, dict[str, float]] = {}
        for _, row in frame.iterrows():
            value = finite_or_none(row[value_col])
            if value is None:
                continue
            by_company.setdefault(int(row["company_id"]), {})[str(row["quarter_end"])] = value
        loaded[input_key] = by_company
    return loaded


def company_inputs(
    all_inputs: dict[str, dict[int, dict[str, float]]],
    company_id: int,
) -> dict[str, dict[str, float]]:
    return {
        input_key: all_inputs[input_key].get(company_id, {})
        for input_key in QUARTERLY_TABLES
    }


def get_component_weights(conn: DbCompat) -> dict[str, float]:
    weights_frame = read_df(
        "SELECT parameter_key, weight FROM business_quarter_trend_weights",
        conn,
    )
    weights = {
        key: float(default_weight)
        for key, _, default_weight, _ in BUSINESS_QUARTER_TREND_DEFAULT_WEIGHTS
    }
    if not weights_frame.empty:
        weights.update({
            str(row["parameter_key"]): float(row["weight"])
            for _, row in weights_frame.dropna(subset=["weight"]).iterrows()
        })
    return weights


def calculate_result(
    source: dict[str, Any],
    all_inputs: dict[str, dict[int, dict[str, float]]],
    component_weights: dict[str, float],
) -> dict[str, Any]:
    result: dict[str, Any] = {
        "Company": source["company"],
        "Ticker": source["ticker"],
    }
    company_id = source.get("company_id")
    if company_id is None:
        result.update({metric: None for metric in METRIC_COLUMNS})
        return result

    details = calculate_business_quarter_trend_details(
        company_inputs(all_inputs, int(company_id)),
        quarter_range=QUARTER_RANGE,
        component_weights=component_weights,
    )
    result.update({
        "Business Quarter Trend Score": finite_or_none(details["business_quarter_trend_score"]),
        "Weighted Median Revenue Growth": finite_or_none(details["weighted_median_revenue_growth"]),
        "Weighted Median Operating Margin": finite_or_none(details["weighted_median_operating_margin"]),
        "Weighted Median Operating Margin Change": finite_or_none(details["weighted_median_operating_margin_change"]),
        "Weighted Median Incremental Operating Margin": finite_or_none(details["weighted_median_incremental_operating_margin"]),
        "Weighted Median Days Sales Outstanding": finite_or_none(details["weighted_median_days_sales_outstanding"]),
        "Weighted Median Capex / OCF": finite_or_none(details["weighted_median_capex_to_ocf"]),
    })
    return result


def style_table(worksheet, table_name: str, widths: dict[str, float], freeze_panes: str) -> None:
    header_fill = PatternFill("solid", fgColor="17365D")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    worksheet.row_dimensions[1].height = 42
    worksheet.freeze_panes = freeze_panes
    worksheet.auto_filter.ref = worksheet.dimensions
    for column, width in widths.items():
        worksheet.column_dimensions[column].width = width
    table = Table(displayName=table_name, ref=worksheet.dimensions)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    worksheet.add_table(table)


def write_workbook(
    output_path: Path,
    source_rows: list[dict[str, Any]],
    result_rows: list[dict[str, Any]],
    component_weights: dict[str, float],
) -> None:
    workbook = Workbook()
    sheet1 = workbook.active
    sheet1.title = "Company Metrics"
    sheet1.append(SHEET1_COLUMNS)
    for result in result_rows:
        sheet1.append([result.get(column) for column in SHEET1_COLUMNS])

    header_to_column = {sheet1.cell(1, column).value: column for column in range(1, sheet1.max_column + 1)}
    for metric in PERCENT_COLUMNS:
        column = header_to_column[metric]
        for row in range(2, sheet1.max_row + 1):
            sheet1.cell(row, column).number_format = "0.00%"
    for row in range(2, sheet1.max_row + 1):
        sheet1.cell(row, header_to_column["Business Quarter Trend Score"]).number_format = "0.0"
        sheet1.cell(row, header_to_column["Weighted Median Days Sales Outstanding"]).number_format = '0.0 "days"'

    unmatched = [source["ticker"] for source in source_rows if source.get("company_id") is None]
    for row_index, (source, result) in enumerate(zip(source_rows, result_rows), start=2):
        if source.get("company_id") is None:
            note = "No matching company record was found in the TaRaSha Research database; research fields are blank."
        elif result["Business Quarter Trend Score"] is None:
            note = "The TaRaSha dashboard could not calculate a score because 16 usable quarters or one or more required component metrics were unavailable."
        else:
            continue
        sheet1.cell(row_index, 1).comment = Comment(note, "Codex")

    weights_text = ", ".join(f"{key}={value:g}" for key, value in component_weights.items())
    sheet1["A1"].comment = Comment(
        "Source: TaRaSha Research, Equity Research > Quarterly Business Trend Score > Dashboard. "
        f"Quarter Range for Score Calculation: {QUARTER_RANGE}. Saved component weights: {weights_text}. "
        "Percentage metrics are stored as decimal values and displayed with Excel percentage formatting. "
        + (f"No database match: {', '.join(unmatched)}." if unmatched else "All source companies matched."),
        "Codex",
    )
    style_table(
        sheet1,
        "QuarterlyTrendCompanyMetrics",
        {"A": 42, "B": 12, "C": 25, "D": 30, "E": 32, "F": 37, "G": 42, "H": 39, "I": 29},
        "C2",
    )

    sheet2 = workbook.create_sheet("Top Quartile Values")
    sheet2.append(["Metric", "Top Quartile Value (75th Percentile)", "Companies with Data"])
    results_frame = pd.DataFrame(result_rows)
    for metric in METRIC_COLUMNS:
        values = pd.to_numeric(results_frame[metric], errors="coerce").replace([np.inf, -np.inf], np.nan).dropna()
        quartile = float(values.quantile(0.75, interpolation="linear")) if not values.empty else None
        sheet2.append([metric, quartile, int(values.count())])

    for row in range(2, sheet2.max_row + 1):
        metric = sheet2.cell(row, 1).value
        if metric in PERCENT_COLUMNS:
            sheet2.cell(row, 2).number_format = "0.00%"
        elif metric == "Business Quarter Trend Score":
            sheet2.cell(row, 2).number_format = "0.0"
        elif metric == "Weighted Median Days Sales Outstanding":
            sheet2.cell(row, 2).number_format = '0.0 "days"'
    sheet2["A1"].comment = Comment(
        "Top quartile is the inclusive 75th percentile (equivalent to Excel QUARTILE.INC(array,3)) "
        "calculated from nonblank Company Metrics values.",
        "Codex",
    )
    style_table(sheet2, "QuarterlyTrendTopQuartile", {"A": 47, "B": 38, "C": 22}, "A2")

    workbook.properties.title = "US-Listed AI Token Cost Beneficiaries - 16-Quarter Business Trend Metrics"
    workbook.properties.subject = "TaRaSha Quarterly Business Trend Dashboard metrics and 75th-percentile cutoffs"
    workbook.properties.creator = "TaRaSha Research"
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("source", type=Path)
    parser.add_argument("output", type=Path)
    args = parser.parse_args()

    source_rows = load_source_companies(args.source)
    session = SessionLocal()
    conn = DbCompat(session)
    try:
        companies = read_df("SELECT id, name, ticker FROM companies ORDER BY id", conn)
        match_company_ids(source_rows, companies)
        company_ids = sorted({int(source["company_id"]) for source in source_rows if source.get("company_id") is not None})
        all_inputs = load_quarterly_inputs(conn, company_ids)
        component_weights = get_component_weights(conn)
        result_rows = [calculate_result(source, all_inputs, component_weights) for source in source_rows]
    finally:
        session.rollback()
        session.close()

    write_workbook(args.output, source_rows, result_rows, component_weights)
    unmatched = [source["ticker"] for source in source_rows if source.get("company_id") is None]
    print(f"Wrote {args.output}")
    print(f"Source rows: {len(source_rows)}; matched: {len(source_rows) - len(unmatched)}; unmatched: {len(unmatched)}")
    if unmatched:
        print("Unmatched tickers: " + ", ".join(unmatched))
    print("Coverage: " + "; ".join(
        f"{metric}={sum(row.get(metric) is not None for row in result_rows)}"
        for metric in METRIC_COLUMNS
    ))


if __name__ == "__main__":
    main()
