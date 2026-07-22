#!/usr/bin/env python3
"""Export Through-the-Cycle Efficiency metrics for the AI beneficiary list."""

from __future__ import annotations

import argparse
import math
import re
import sys
from pathlib import Path
from typing import Any, Optional

PROJECT_ROOT = Path(__file__).resolve().parents[1]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

import numpy as np
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo

from core_backend import calculate_scenario_incremental_operating_margin, read_df
from db_session import DbCompat, SessionLocal
from ttc_efficiency import (
    _build_cash_and_equivalents_series,
    _capital_intensity,
    _free_cash_flow,
    _median,
    _normalize_label,
    _score_balance_sheet,
    _score_cash_flow,
    _score_income_statement,
    _stdev_sample,
)


YEAR_RANGE_END = 2020

SHEET1_COLUMNS = [
    "Company",
    "Ticker",
    "Income Statement Efficiency Score (0-100)",
    "Incremental Margin % [Pricing Power + Operating Leverage]",
    "Balance Sheet Strength Score",
    "Net Debt/EBITDA",
    "Interest Coverage",
    "Quick Ratio",
    "Current Ratio",
    "Debt to Capitalization",
    "Debt Maturity Pressure",
    "Cash Flow Efficiency Score (0-100)",
    "Operating Cash Flow Margin",
    "FCF Margin",
    "FCF Margin Standard Deviation",
    "CFO/Net Income",
    "Cash Return on Invested Capital (CROIC)",
    "Capital intensity (Capex/Revenue)",
]

METRIC_COLUMNS = SHEET1_COLUMNS[2:]

HIGHER_IS_BETTER = {
    "Income Statement Efficiency Score (0-100)",
    "Incremental Margin % [Pricing Power + Operating Leverage]",
    "Balance Sheet Strength Score",
    "Interest Coverage",
    "Cash Flow Efficiency Score (0-100)",
    "Operating Cash Flow Margin",
    "FCF Margin",
    "CFO/Net Income",
    "Cash Return on Invested Capital (CROIC)",
}

LOWER_IS_BETTER = set(METRIC_COLUMNS) - HIGHER_IS_BETTER

PERCENT_COLUMNS = {
    "Incremental Margin % [Pricing Power + Operating Leverage]",
    "Operating Cash Flow Margin",
    "FCF Margin",
    "FCF Margin Standard Deviation",
    "CFO/Net Income",
    "Cash Return on Invested Capital (CROIC)",
    "Capital intensity (Capex/Revenue)",
}

SCORE_COLUMNS = {
    "Income Statement Efficiency Score (0-100)",
    "Balance Sheet Strength Score",
    "Cash Flow Efficiency Score (0-100)",
}

ANNUAL_TABLES = {
    "revenue": ("revenues_annual", "revenue"),
    "cost_of_revenue": ("cost_of_revenue_annual", "cost_of_revenue"),
    "sga": ("sga_annual", "sga"),
    "operating_income": ("operating_income_annual", "operating_income"),
    "ebit": ("ebit_annual", "ebit"),
    "total_debt": ("total_debt_annual", "total_debt"),
    "cash": ("cash_and_cash_equivalents_annual", "cash_and_cash_equivalents"),
    "accounts_receivable": ("accounts_receivable_annual", "accounts_receivable"),
    "current_assets": ("total_current_assets_annual", "total_current_assets"),
    "current_liabilities": ("total_current_liabilities_annual", "total_current_liabilities"),
    "current_debt": ("current_debt_annual", "current_debt"),
    "shareholders_equity": ("shareholders_equity_annual", "shareholders_equity"),
    "ebitda": ("ebitda_annual", "ebitda"),
    "interest_expense": ("interest_expense_annual", "interest_expense"),
    "operating_cash_flow": ("operating_cash_flow_annual", "operating_cash_flow"),
    "capital_expenditures": ("capital_expenditures_annual", "capital_expenditures"),
    "net_income": ("net_income_annual", "net_income"),
    "net_ppe": ("net_ppe_annual", "net_ppe"),
    "short_term_investments": ("short_term_investments_annual", "short_term_investments"),
    "goodwill_and_intangibles": ("goodwill_and_intangibles_annual", "goodwill_and_intangibles"),
    "other_long_term_assets": ("other_long_term_assets_annual", "other_long_term_assets"),
    "deferred_revenue": ("deferred_revenue_annual", "deferred_revenue"),
    "deferred_tax_liabilities": ("deferred_tax_liabilities_annual", "deferred_tax_liabilities"),
    "other_long_term_liabilities": ("other_long_term_liabilities_annual", "other_long_term_liabilities"),
}

TTM_TABLES = {
    "revenue": ("revenues_ttm", "revenue"),
    "operating_cash_flow": ("operating_cash_flow_ttm", "operating_cash_flow"),
    "capital_expenditures": ("capital_expenditures_ttm", "capital_expenditures"),
    "net_income": ("net_income_ttm", "net_income"),
    "net_ppe": ("net_ppe_ttm", "net_ppe"),
    "current_assets": ("total_current_assets_ttm", "total_current_assets"),
    "current_liabilities": ("total_current_liabilities_ttm", "total_current_liabilities"),
    "current_debt": ("current_debt_ttm", "current_debt"),
    "cash": ("cash_and_cash_equivalents_ttm", "cash_and_cash_equivalents"),
    "short_term_investments": ("short_term_investments_ttm", "short_term_investments"),
    "goodwill_and_intangibles": ("goodwill_and_intangibles_ttm", "goodwill_and_intangibles"),
    "other_long_term_assets": ("other_long_term_assets_ttm", "other_long_term_assets"),
    "deferred_revenue": ("deferred_revenue_ttm", "deferred_revenue"),
    "deferred_tax_liabilities": ("deferred_tax_liabilities_ttm", "deferred_tax_liabilities"),
    "other_long_term_liabilities": ("other_long_term_liabilities_ttm", "other_long_term_liabilities"),
}


def normalized_ticker(value: Any) -> str:
    return str(value or "").strip().upper().replace(".", "-")


def normalized_name(value: Any) -> str:
    return re.sub(r"[^a-z0-9]+", "", str(value or "").lower())


def finite_or_none(value: Any) -> Optional[float]:
    if value is None:
        return None
    try:
        number = float(value)
    except (TypeError, ValueError):
        return None
    return number if math.isfinite(number) else None


def load_source_companies(source_path: Path) -> list[dict[str, Any]]:
    workbook = load_workbook(source_path, read_only=True, data_only=True)
    worksheet = workbook["Companies"]
    rows: list[dict[str, Any]] = []
    for values in worksheet.iter_rows(min_row=6, values_only=True):
        rank, company, ticker = values[:3]
        if rank is None:
            continue
        rows.append({"rank": int(rank), "company": str(company).strip(), "ticker": str(ticker).strip()})
    workbook.close()
    return rows


def match_company_ids(source_rows: list[dict[str, Any]], companies: pd.DataFrame) -> None:
    companies = companies.copy()
    companies["ticker_norm"] = companies["ticker"].map(normalized_ticker)
    companies["name_norm"] = companies["name"].map(normalized_name)
    for source in source_rows:
        candidates = companies[companies["ticker_norm"] == normalized_ticker(source["ticker"])]
        if len(candidates) > 1:
            exact = candidates[candidates["name_norm"] == normalized_name(source["company"])]
            if len(exact) == 1:
                candidates = exact
        source["company_id"] = int(candidates.iloc[0]["id"]) if len(candidates) == 1 else None


def load_annual_data(conn: DbCompat, company_ids: list[int]) -> dict[str, dict[int, dict[int, float]]]:
    placeholders = ",".join(["?"] * len(company_ids))
    loaded: dict[str, dict[int, dict[int, float]]] = {}
    for key, (table, value_col) in ANNUAL_TABLES.items():
        frame = read_df(
            f"""
            SELECT company_id, fiscal_year AS year, {value_col}
            FROM {table}
            WHERE company_id IN ({placeholders})
            ORDER BY company_id, fiscal_year
            """,
            conn,
            params=company_ids,
        )
        by_company: dict[int, dict[int, float]] = {}
        for _, row in frame.iterrows():
            value = finite_or_none(row[value_col])
            if value is not None:
                by_company.setdefault(int(row["company_id"]), {})[int(row["year"])] = value
        loaded[key] = by_company
    return loaded


def load_ttm_data(conn: DbCompat, company_ids: list[int]) -> dict[str, dict[int, dict[int, float]]]:
    placeholders = ",".join(["?"] * len(company_ids))
    loaded: dict[str, dict[int, dict[int, float]]] = {}
    for key, (table, value_col) in TTM_TABLES.items():
        try:
            frame = read_df(
                f"""
                SELECT company_id, as_of, {value_col}
                FROM {table}
                WHERE company_id IN ({placeholders})
                ORDER BY company_id, as_of
                """,
                conn,
                params=company_ids,
            )
        except Exception:
            loaded[key] = {}
            continue
        by_company: dict[int, dict[int, float]] = {}
        for _, row in frame.iterrows():
            value = finite_or_none(row[value_col])
            as_of = str(row.get("as_of", "")).strip()
            if value is None or len(as_of) < 4:
                continue
            try:
                year = int(as_of[:4])
            except ValueError:
                continue
            by_company.setdefault(int(row["company_id"]), {})[year] = value
        loaded[key] = by_company
    return loaded


def series_for(
    annual_data: dict[str, dict[int, dict[int, float]]],
    key: str,
    company_id: int,
) -> dict[int, float]:
    return dict(annual_data.get(key, {}).get(company_id, {}))


def cash_flow_series_for(
    annual_data: dict[str, dict[int, dict[int, float]]],
    ttm_data: dict[str, dict[int, dict[int, float]]],
    key: str,
    company_id: int,
) -> dict[int, float]:
    output = series_for(annual_data, key, company_id)
    output.update(ttm_data.get(key, {}).get(company_id, {}))
    return output


def parse_assumption_params(assumptions: pd.DataFrame) -> tuple[dict, dict, dict]:
    sections: dict[str, list[dict[str, Any]]] = {}
    for _, row in assumptions.iterrows():
        sections.setdefault(str(row["section"]), []).append({
            "metric": str(row["metric"]),
            "weight": float(row["weight"]),
            "threshold": float(row["threshold"]),
        })

    income = {
        "operating_margin": (0.0, 0.0),
        "gross_margin": (0.0, 0.0),
        "sga_ratio": (0.0, 0.0),
        "incremental_margin": (0.0, 0.0),
        "op_margin_volatility": (0.0, 0.0),
    }
    for row in sections.get("Income Statement Efficiency Score", []):
        normalized = _normalize_label(row["metric"])
        pair = (row["weight"], row["threshold"])
        if "volatility" in normalized or "stddev" in normalized:
            income["op_margin_volatility"] = pair
        elif "operatingmargin" in normalized:
            income["operating_margin"] = pair
        elif "grossmargin" in normalized:
            income["gross_margin"] = pair
        elif "sga" in normalized or "sellinggeneral" in normalized:
            income["sga_ratio"] = pair
        elif "incrementalmargin" in normalized:
            income["incremental_margin"] = pair

    balance = {
        "net_debt_ebitda": (0.0, 0.0),
        "interest_coverage": (0.0, 0.0),
        "quick_ratio": (0.0, 0.0),
        "current_ratio": (0.0, 0.0),
        "debt_to_capitalization": (0.0, 0.0),
        "debt_maturity_pressure": (0.0, 0.0),
    }
    for row in sections.get("Balance Sheet Strength Score", []):
        normalized = _normalize_label(row["metric"])
        pair = (row["weight"], row["threshold"])
        if "netdebt" in normalized and "ebitda" in normalized:
            balance["net_debt_ebitda"] = pair
        elif "interestcoverage" in normalized:
            balance["interest_coverage"] = pair
        elif "quickratio" in normalized:
            balance["quick_ratio"] = pair
        elif "currentratio" in normalized:
            balance["current_ratio"] = pair
        elif "debt" in normalized and "capital" in normalized:
            balance["debt_to_capitalization"] = pair
        elif "maturity" in normalized:
            balance["debt_maturity_pressure"] = pair

    cash_flow = {
        "ocf_margin": (0.0, 0.0),
        "fcf_margin": (0.0, 0.0),
        "cfo_net_income": (0.0, 0.0),
        "croic": (0.0, 0.0),
        "capital_intensity": (0.0, 0.0),
        "fcf_margin_volatility": (0.0, 0.0),
    }
    for row in sections.get("Cash Flow Efficiency Score", []):
        normalized = _normalize_label(row["metric"])
        pair = (row["weight"], row["threshold"])
        if "operatingcashflowmargin" in normalized or ("cfo" in normalized and "revenue" in normalized and "margin" in normalized):
            cash_flow["ocf_margin"] = pair
        elif "freecashflowmargin" in normalized and "volatility" not in normalized and "stddev" not in normalized:
            cash_flow["fcf_margin"] = pair
        elif "cfo" in normalized and "netincome" in normalized:
            cash_flow["cfo_net_income"] = pair
        elif "croic" in normalized or ("cashreturn" in normalized and "investedcapital" in normalized):
            cash_flow["croic"] = pair
        elif "capitalintensity" in normalized or ("capex" in normalized and "revenue" in normalized):
            cash_flow["capital_intensity"] = pair
        elif ("fcfmargin" in normalized and "volatility" in normalized) or ("fcfmargin" in normalized and "stddev" in normalized):
            cash_flow["fcf_margin_volatility"] = pair

    return income, balance, cash_flow


def income_metrics(
    company_id: int,
    annual_data: dict[str, dict[int, dict[int, float]]],
    params: dict,
) -> dict[str, Optional[float]]:
    revenue = series_for(annual_data, "revenue", company_id)
    if not revenue:
        return {
            "Income Statement Efficiency Score (0-100)": None,
            "Incremental Margin % [Pricing Power + Operating Leverage]": None,
        }
    cogs = series_for(annual_data, "cost_of_revenue", company_id)
    sga = series_for(annual_data, "sga", company_id)
    operating_income = series_for(annual_data, "operating_income", company_id)
    if not operating_income:
        operating_income = series_for(annual_data, "ebit", company_id)

    years = [year for year in sorted(revenue) if YEAR_RANGE_END <= year <= max(revenue)]
    op_margin_values: list[float] = []
    gross_margin_values: list[float] = []
    sga_ratio_values: list[float] = []
    incremental_values: list[float] = []
    for year in years:
        rev = revenue.get(year)
        oi = operating_income.get(year)
        if rev is not None and rev != 0 and oi is not None:
            op_margin_values.append(oi / rev)
        cost = cogs.get(year)
        if rev is not None and rev != 0 and cost is not None:
            gross_margin_values.append((rev - cost) / rev)
        sga_value = sga.get(year)
        if rev is not None and rev != 0 and sga_value is not None:
            sga_ratio_values.append(sga_value / rev)
        if year - 1 in revenue and year - 1 in operating_income and rev is not None and oi is not None:
            prior_rev = revenue.get(year - 1)
            prior_oi = operating_income.get(year - 1)
            if prior_rev is not None and prior_oi is not None:
                incremental = calculate_scenario_incremental_operating_margin(rev, oi, prior_rev, prior_oi)["value"]
                if incremental is not None:
                    incremental_values.append(float(incremental))

    op_margin_median = _median(op_margin_values)
    op_margin_stdev = _stdev_sample(op_margin_values)
    gross_margin_median = _median(gross_margin_values)
    sga_ratio_median = _median(sga_ratio_values)
    incremental_median = _median(incremental_values)
    score = _score_income_statement(
        op_margin_median,
        op_margin_stdev,
        gross_margin_median,
        sga_ratio_median,
        incremental_median,
        params,
    )
    return {
        "Income Statement Efficiency Score (0-100)": finite_or_none(score),
        "Incremental Margin % [Pricing Power + Operating Leverage]": finite_or_none(incremental_median),
    }


def balance_metrics(
    company_id: int,
    annual_data: dict[str, dict[int, dict[int, float]]],
    params: dict,
) -> dict[str, Optional[float]]:
    total_debt = series_for(annual_data, "total_debt", company_id)
    empty = {
        "Balance Sheet Strength Score": None,
        "Net Debt/EBITDA": None,
        "Interest Coverage": None,
        "Quick Ratio": None,
        "Current Ratio": None,
        "Debt to Capitalization": None,
        "Debt Maturity Pressure": None,
    }
    if not total_debt:
        return empty

    cash = series_for(annual_data, "cash", company_id)
    accounts_receivable = series_for(annual_data, "accounts_receivable", company_id)
    current_assets = series_for(annual_data, "current_assets", company_id)
    current_liabilities = series_for(annual_data, "current_liabilities", company_id)
    current_debt = series_for(annual_data, "current_debt", company_id)
    equity = series_for(annual_data, "shareholders_equity", company_id)
    ebitda = series_for(annual_data, "ebitda", company_id)
    operating_income = series_for(annual_data, "operating_income", company_id)
    interest_expense = series_for(annual_data, "interest_expense", company_id)

    available_years = sorted(set(total_debt) | set(cash) | set(current_assets) | set(current_liabilities))
    if not available_years:
        return empty
    years = [year for year in available_years if YEAR_RANGE_END <= year <= max(available_years)]

    nd_ebitda_values: list[float] = []
    interest_coverage_values: list[float] = []
    quick_ratio_values: list[float] = []
    current_ratio_values: list[float] = []
    debt_to_cap_values: list[float] = []
    maturity_values: list[float] = []
    nd_penalty = False

    for year in years:
        debt = total_debt.get(year)
        if debt is not None:
            cash_value = cash.get(year, 0.0)
            ebitda_value = ebitda.get(year)
            net_debt = debt - cash_value
            if ebitda_value is not None and ebitda_value <= 0 and net_debt > 0:
                nd_penalty = True
            if ebitda_value is not None and ebitda_value != 0:
                nd_ebitda_values.append(net_debt / ebitda_value)
            equity_value = equity.get(year)
            if equity_value is not None and debt + equity_value != 0:
                debt_to_cap_values.append(debt / (debt + equity_value))
            current_debt_value = current_debt.get(year)
            if current_debt_value is not None:
                maturity_values.append(0.0 if debt == 0 or current_debt_value == 0 else current_debt_value / debt)

        oi = operating_income.get(year)
        interest = interest_expense.get(year)
        if oi is not None and interest is not None:
            interest_coverage_values.append(oi / interest if interest != 0 else (100.0 if oi > 0 else 0.0))
        current_assets_value = current_assets.get(year)
        current_liabilities_value = current_liabilities.get(year)
        if current_assets_value is not None and current_liabilities_value is not None and current_liabilities_value != 0:
            current_ratio_values.append(current_assets_value / current_liabilities_value)
        if current_liabilities_value is not None and current_liabilities_value != 0:
            quick_ratio_values.append((cash.get(year, 0.0) + accounts_receivable.get(year, 0.0)) / current_liabilities_value)

    nd_ebitda_median = _median(nd_ebitda_values)
    interest_coverage_median = _median(interest_coverage_values)
    quick_ratio_median = _median(quick_ratio_values)
    current_ratio_median = _median(current_ratio_values)
    debt_to_cap_median = _median(debt_to_cap_values)
    maturity_median = _median(maturity_values)
    score = _score_balance_sheet(
        nd_ebitda_median,
        interest_coverage_median,
        quick_ratio_median,
        current_ratio_median,
        debt_to_cap_median,
        maturity_median,
        nd_penalty,
        params,
    )
    return {
        "Balance Sheet Strength Score": finite_or_none(score),
        "Net Debt/EBITDA": finite_or_none(nd_ebitda_median),
        "Interest Coverage": finite_or_none(interest_coverage_median),
        "Quick Ratio": finite_or_none(quick_ratio_median),
        "Current Ratio": finite_or_none(current_ratio_median),
        "Debt to Capitalization": finite_or_none(debt_to_cap_median),
        "Debt Maturity Pressure": finite_or_none(maturity_median),
    }


def cash_flow_metrics(
    company_id: int,
    annual_data: dict[str, dict[int, dict[int, float]]],
    ttm_data: dict[str, dict[int, dict[int, float]]],
    params: dict,
) -> dict[str, Optional[float]]:
    keys = list(TTM_TABLES)
    series = {key: cash_flow_series_for(annual_data, ttm_data, key, company_id) for key in keys}
    empty = {
        "Cash Flow Efficiency Score (0-100)": None,
        "Operating Cash Flow Margin": None,
        "FCF Margin": None,
        "FCF Margin Standard Deviation": None,
        "CFO/Net Income": None,
        "Cash Return on Invested Capital (CROIC)": None,
        "Capital intensity (Capex/Revenue)": None,
    }
    if not series["revenue"]:
        return empty

    cash_ex_sti = _build_cash_and_equivalents_series(series["cash"], series["short_term_investments"])
    available_years = sorted(set().union(*(set(values) for values in series.values()), set(cash_ex_sti)))
    if not available_years:
        return empty
    years = [year for year in available_years if YEAR_RANGE_END <= year <= max(available_years)]
    if not years:
        return empty

    ocf_margin_values: list[float] = []
    fcf_margin_values: list[float] = []
    cfo_net_income_values: list[float] = []
    croic_values: list[float] = []
    capital_intensity_values: list[float] = []
    for year in years:
        revenue = series["revenue"].get(year)
        cfo = series["operating_cash_flow"].get(year)
        capex = series["capital_expenditures"].get(year)
        net_income = series["net_income"].get(year)
        capex_outflow = abs(capex) if capex is not None else None
        free_cash_flow = _free_cash_flow(cfo, capex_outflow) if cfo is not None and capex_outflow is not None else None

        if revenue is not None and revenue != 0 and cfo is not None:
            ocf_margin_values.append(cfo / revenue)
        if revenue is not None and revenue != 0 and free_cash_flow is not None:
            fcf_margin_values.append(free_cash_flow / revenue)
        if cfo is not None and net_income is not None and net_income != 0:
            cfo_net_income_values.append(cfo / net_income)
        if revenue is not None and revenue != 0 and capex_outflow is not None:
            capital_intensity_values.append(_capital_intensity(capex_outflow, revenue))

        required = [
            free_cash_flow,
            series["net_ppe"].get(year),
            series["current_assets"].get(year),
            series["current_liabilities"].get(year),
            series["current_debt"].get(year),
            cash_ex_sti.get(year),
            series["goodwill_and_intangibles"].get(year),
            series["other_long_term_assets"].get(year),
            series["deferred_revenue"].get(year),
            series["other_long_term_liabilities"].get(year),
        ]
        if all(value is not None for value in required):
            sti = series["short_term_investments"].get(year, 0.0)
            deferred_tax = series["deferred_tax_liabilities"].get(year, 0.0)
            current_operating_assets = series["current_assets"][year] - cash_ex_sti[year] - sti
            current_operating_liabilities = series["current_liabilities"][year] - series["current_debt"][year]
            net_working_capital = current_operating_assets - current_operating_liabilities
            other_operating_assets = series["goodwill_and_intangibles"][year] + series["other_long_term_assets"][year]
            non_interest_operating_liabilities = series["deferred_revenue"][year] + deferred_tax + series["other_long_term_liabilities"][year]
            invested_capital = series["net_ppe"][year] + net_working_capital + other_operating_assets - non_interest_operating_liabilities
            if invested_capital != 0:
                croic_values.append(free_cash_flow / invested_capital)

    ocf_margin_median = _median(ocf_margin_values)
    fcf_margin_median = _median(fcf_margin_values)
    fcf_margin_stdev = _stdev_sample(fcf_margin_values)
    cfo_net_income_median = _median(cfo_net_income_values)
    croic_median = _median(croic_values)
    capital_intensity_median = _median(capital_intensity_values)
    score = _score_cash_flow(
        ocf_margin_median,
        fcf_margin_median,
        cfo_net_income_median,
        croic_median,
        capital_intensity_median,
        fcf_margin_stdev,
        params,
    )
    return {
        "Cash Flow Efficiency Score (0-100)": finite_or_none(score),
        "Operating Cash Flow Margin": finite_or_none(ocf_margin_median),
        "FCF Margin": finite_or_none(fcf_margin_median),
        "FCF Margin Standard Deviation": finite_or_none(fcf_margin_stdev),
        "CFO/Net Income": finite_or_none(cfo_net_income_median),
        "Cash Return on Invested Capital (CROIC)": finite_or_none(croic_median),
        "Capital intensity (Capex/Revenue)": finite_or_none(capital_intensity_median),
    }


def style_table(worksheet, table_name: str, widths: dict[str, float], freeze_panes: str) -> None:
    fill = PatternFill("solid", fgColor="17365D")
    font = Font(color="FFFFFF", bold=True)
    for cell in worksheet[1]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    worksheet.row_dimensions[1].height = 48
    worksheet.freeze_panes = freeze_panes
    worksheet.auto_filter.ref = worksheet.dimensions
    for column, width in widths.items():
        worksheet.column_dimensions[column].width = width
    table = Table(displayName=table_name, ref=worksheet.dimensions)
    table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
    worksheet.add_table(table)


def write_workbook(
    output_path: Path,
    source_rows: list[dict[str, Any]],
    result_rows: list[dict[str, Any]],
) -> None:
    workbook = Workbook()
    sheet1 = workbook.active
    sheet1.title = "Company Metrics"
    sheet1.append(SHEET1_COLUMNS)
    for result in result_rows:
        sheet1.append([result.get(column) for column in SHEET1_COLUMNS])

    header_columns = {sheet1.cell(1, column).value: column for column in range(1, sheet1.max_column + 1)}
    for metric in METRIC_COLUMNS:
        column = header_columns[metric]
        number_format = "0.00%" if metric in PERCENT_COLUMNS else ("0.0" if metric in SCORE_COLUMNS else "0.00")
        for row in range(2, sheet1.max_row + 1):
            sheet1.cell(row, column).number_format = number_format

    unmatched = [source["ticker"] for source in source_rows if source.get("company_id") is None]
    for row_index, (source, result) in enumerate(zip(source_rows, result_rows), start=2):
        if source.get("company_id") is None:
            note = "No matching company record was found in the TaRaSha Research database; research fields are blank."
        elif all(result.get(metric) is None for metric in METRIC_COLUMNS):
            note = "No Through-the-Cycle Efficiency metrics were available for the requested year range."
        else:
            continue
        sheet1.cell(row_index, 1).comment = Comment(note, "Codex")

    sheet1["A1"].comment = Comment(
        "Source: TaRaSha Research, Equity Research > Through-the-Cycle Efficiency Score. "
        "Year range: Recent - 2020. Percentage metrics are stored as decimal values and displayed with Excel percentage formatting. "
        + (f"No database match: {', '.join(unmatched)}." if unmatched else "All source companies matched."),
        "Codex",
    )
    widths = {"A": 42, "B": 12}
    for column in "CDEFGHIJKLMNOPQR":
        widths[column] = 25
    widths["D"] = 48
    widths["Q"] = 39
    widths["R"] = 34
    style_table(sheet1, "TTCEfficiencyCompanyMetrics", widths, "C2")

    sheet2 = workbook.create_sheet("Top Quartile Values")
    sheet2.append(["Metric", "Better Direction", "Top Quartile Cutoff", "Percentile Used", "Companies with Data"])
    result_frame = pd.DataFrame(result_rows)
    for metric in METRIC_COLUMNS:
        values = pd.to_numeric(result_frame[metric], errors="coerce").replace([np.inf, -np.inf], np.nan).dropna()
        higher = metric in HIGHER_IS_BETTER
        percentile = 0.75 if higher else 0.25
        cutoff = float(values.quantile(percentile, interpolation="linear")) if not values.empty else None
        sheet2.append([
            metric,
            "Higher is better" if higher else "Lower is better",
            cutoff,
            "75th percentile" if higher else "25th percentile",
            int(values.count()),
        ])
    for row in range(2, sheet2.max_row + 1):
        metric = sheet2.cell(row, 1).value
        sheet2.cell(row, 3).number_format = "0.00%" if metric in PERCENT_COLUMNS else ("0.0" if metric in SCORE_COLUMNS else "0.00")
    sheet2["A1"].comment = Comment(
        "Performance-adjusted top-quartile cutoffs use the inclusive 75th percentile for higher-is-better metrics "
        "and the inclusive 25th percentile for lower-is-better metrics.",
        "Codex",
    )
    style_table(
        sheet2,
        "TTCEfficiencyTopQuartile",
        {"A": 54, "B": 20, "C": 24, "D": 20, "E": 22},
        "A2",
    )

    workbook.properties.title = "US-Listed AI Token Cost Beneficiaries - Through-the-Cycle Efficiency Metrics"
    workbook.properties.subject = "TaRaSha TTC efficiency metrics and performance-adjusted top-quartile cutoffs"
    workbook.properties.creator = "TaRaSha Research"
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("source", type=Path)
    parser.add_argument("output", type=Path)
    args = parser.parse_args()

    source_rows = load_source_companies(args.source)
    session = SessionLocal()
    conn = DbCompat(session)
    try:
        companies = read_df("SELECT id, name, ticker FROM companies ORDER BY id", conn)
        match_company_ids(source_rows, companies)
        company_ids = sorted({int(source["company_id"]) for source in source_rows if source.get("company_id") is not None})
        annual_data = load_annual_data(conn, company_ids)
        ttm_data = load_ttm_data(conn, company_ids)
        assumptions = read_df(
            """
            SELECT section, metric, weight, threshold
            FROM ttc_assumptions
            WHERE section IN (?, ?, ?)
            ORDER BY section, sort_order, id
            """,
            conn,
            params=[
                "Income Statement Efficiency Score",
                "Balance Sheet Strength Score",
                "Cash Flow Efficiency Score",
            ],
        )
        income_params, balance_params, cash_flow_params = parse_assumption_params(assumptions)

        result_rows: list[dict[str, Any]] = []
        for source in source_rows:
            result: dict[str, Any] = {"Company": source["company"], "Ticker": source["ticker"]}
            company_id = source.get("company_id")
            if company_id is None:
                result.update({metric: None for metric in METRIC_COLUMNS})
            else:
                result.update(income_metrics(int(company_id), annual_data, income_params))
                result.update(balance_metrics(int(company_id), annual_data, balance_params))
                result.update(cash_flow_metrics(int(company_id), annual_data, ttm_data, cash_flow_params))
            result_rows.append(result)
    finally:
        session.rollback()
        session.close()

    write_workbook(args.output, source_rows, result_rows)
    unmatched = [source["ticker"] for source in source_rows if source.get("company_id") is None]
    print(f"Wrote {args.output}")
    print(f"Source rows: {len(source_rows)}; matched: {len(source_rows) - len(unmatched)}; unmatched: {len(unmatched)}")
    if unmatched:
        print("Unmatched tickers: " + ", ".join(unmatched))
    print("Coverage: " + "; ".join(
        f"{metric}={sum(row.get(metric) is not None for row in result_rows)}"
        for metric in METRIC_COLUMNS
    ))


if __name__ == "__main__":
    main()
