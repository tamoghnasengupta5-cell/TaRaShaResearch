#!/usr/bin/env python3
"""Export Combined Dashboard breakup metrics for the AI token beneficiary list.

The calculations mirror ``combined_dashboard.py`` with its default settings:
``Recent - 2020`` and sample standard deviation (ddof=1).  The script is
read-only with respect to the TaRaSha database; it uses the currently stored
annual and derived research series.
"""

from __future__ import annotations

import argparse
import math
import re
import sys
from pathlib import Path
from typing import Any, Iterable, Optional

PROJECT_ROOT = Path(__file__).resolve().parents[1]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

import numpy as np
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo

from core_backend import (
    compute_growth_stats,
    compute_margin_growth_stats,
    compute_margin_stats,
    exclude_recent_zero_accumulated_profit_for_stats,
    read_df,
)
from db_session import DbCompat, SessionLocal


SHEET1_COLUMNS = [
    "Company",
    "Ticker",
    "Debt-Adj. Total Score",
    "Median ROIC-WACC",
    "ROIC-WACC Std Dev",
    "Median Op. Margin",
    "Op. Margin Std Dev",
    "Median Op. Margin Growth",
    "Op. Margin Growth Std Dev",
    "Median Revenue Growth",
    "Median ROE",
    "ROE Std Dev",
    "Median FCFF Change",
]

METRIC_COLUMNS = SHEET1_COLUMNS[2:]

TABLE_SPECS = {
    "acc": ("accumulated_profit_annual", "accumulated_profit"),
    "roe": ("roe_annual", "roe"),
    "roce": ("roce_annual", "roce"),
    "interest_load": ("interest_load_annual", "interest_load_pct"),
    "revenue": ("revenues_annual", "revenue"),
    "pretax": ("pretax_income_annual", "pretax_income"),
    "net_income": ("net_income_annual", "net_income"),
    "nopat": ("nopat_annual", "nopat"),
    "op_margin": ("op_margin_annual", "margin"),
    "fcff": ("fcff_annual", "fcff"),
    "spread": ("roic_wacc_spread_annual", "spread_pct"),
}


def normalized_ticker(value: Any) -> str:
    return str(value or "").strip().upper().replace(".", "-")


def normalized_name(value: Any) -> str:
    return re.sub(r"[^a-z0-9]+", "", str(value or "").lower())


def finite_or_none(value: Any) -> Optional[float]:
    if value is None:
        return None
    try:
        number = float(value)
    except (TypeError, ValueError):
        return None
    return number if math.isfinite(number) else None


def weighted_score(pairs: Iterable[tuple[Optional[float], Optional[float]]]) -> Optional[float]:
    numerator = 0.0
    denominator = 0.0
    for value, weight in pairs:
        if value is not None and weight is not None and weight > 0:
            numerator += value * weight
            denominator += weight
    if denominator == 0.0:
        return None
    return numerator / denominator


def get_weight(weight_map: dict[str, float], *names: str) -> Optional[float]:
    for name in names:
        if name in weight_map:
            return weight_map[name]
    return None


def level_stats(
    df: pd.DataFrame,
    yr_start: int,
    yr_end: int,
    value_col: str,
) -> tuple[Optional[float], Optional[float]]:
    if df.empty:
        return None, None
    ranged = df[(df["year"] <= yr_start) & (df["year"] >= yr_end)]
    values = pd.to_numeric(ranged[value_col], errors="coerce").dropna().to_numpy(dtype=float)
    if values.size == 0:
        return None, None
    median = float(np.median(values))
    stdev = 0.0 if values.size < 2 else float(np.std(values, ddof=1))
    return median, stdev


def grouped_frame(all_rows: pd.DataFrame, company_id: Optional[int]) -> pd.DataFrame:
    if company_id is None or all_rows.empty:
        return all_rows.iloc[0:0].drop(columns=["company_id"], errors="ignore").copy()
    result = all_rows[all_rows["company_id"] == company_id].copy()
    return result.drop(columns=["company_id"], errors="ignore").sort_values("year").reset_index(drop=True)


def load_source_companies(source_path: Path) -> list[dict[str, Any]]:
    workbook = load_workbook(source_path, read_only=True, data_only=True)
    worksheet = workbook["Companies"]
    rows: list[dict[str, Any]] = []
    for values in worksheet.iter_rows(min_row=6, values_only=True):
        rank, company, ticker = values[:3]
        if rank is None:
            continue
        rows.append({"rank": int(rank), "company": str(company).strip(), "ticker": str(ticker).strip()})
    workbook.close()
    return rows


def match_company_ids(source_rows: list[dict[str, Any]], companies: pd.DataFrame) -> None:
    companies = companies.copy()
    companies["ticker_norm"] = companies["ticker"].map(normalized_ticker)
    companies["name_norm"] = companies["name"].map(normalized_name)

    for source in source_rows:
        candidates = companies[companies["ticker_norm"] == normalized_ticker(source["ticker"])]
        source_name = normalized_name(source["company"])
        if len(candidates) > 1:
            exact = candidates[candidates["name_norm"] == source_name]
            if len(exact) == 1:
                candidates = exact
        if len(candidates) == 1:
            source["company_id"] = int(candidates.iloc[0]["id"])
            source["database_name"] = str(candidates.iloc[0]["name"])
        else:
            source["company_id"] = None
            source["database_name"] = None


def load_annual_tables(conn: DbCompat, company_ids: list[int]) -> dict[str, pd.DataFrame]:
    placeholders = ",".join(["?"] * len(company_ids))
    frames: dict[str, pd.DataFrame] = {}
    for key, (table, value_col) in TABLE_SPECS.items():
        frames[key] = read_df(
            f"""
            SELECT company_id, fiscal_year AS year, {value_col}
            FROM {table}
            WHERE company_id IN ({placeholders})
            ORDER BY company_id, fiscal_year
            """,
            conn,
            params=company_ids,
        )
    return frames


def compute_company_metrics(
    company_id: int,
    frames: dict[str, pd.DataFrame],
    growth_weights: dict[str, float],
    stddev_weights: dict[str, float],
) -> dict[str, Optional[float]]:
    series = {name: grouped_frame(frame, company_id) for name, frame in frames.items()}

    median_roe = std_roe = None
    median_rev = None
    median_om = std_om = None
    median_om_growth = std_om_growth = None
    median_fcff = None
    median_spread = std_spread = None
    bs_debt = pl_scaled = fs_scaled = None

    # Balance-sheet component.  The range starts at the most recent available
    # annual year among the dashboard's balance-sheet series and ends at 2020.
    bs_years: list[int] = []
    for key in ("acc", "roe", "roce", "interest_load"):
        if not series[key].empty:
            bs_years.extend(pd.to_numeric(series[key]["year"], errors="coerce").dropna().astype(int).tolist())
    if bs_years:
        yr_start, yr_end = max(bs_years), 2020
        acc_stats = exclude_recent_zero_accumulated_profit_for_stats(series["acc"])
        med_acc, std_acc = compute_growth_stats(
            acc_stats,
            yr_start,
            yr_end,
            stdev_sample=True,
            value_col="accumulated_profit",
            abs_denom=True,
        ) if not acc_stats.empty else (None, None)

        med_roe_raw, std_roe_raw = level_stats(series["roe"], yr_start, yr_end, "roe")
        med_roce_raw, std_roce_raw = level_stats(series["roce"], yr_start, yr_end, "roce")
        med_interest, _ = level_stats(series["interest_load"], yr_start, yr_end, "interest_load_pct")

        med_acc_pct = None if med_acc is None else med_acc * 100.0
        std_acc_pct = None if std_acc is None else std_acc * 100.0
        median_roe = None if med_roe_raw is None else med_roe_raw * 100.0
        std_roe = None if std_roe_raw is None else std_roe_raw * 100.0
        med_roce_pct = None if med_roce_raw is None else med_roce_raw * 100.0
        std_roce_pct = None if std_roce_raw is None else std_roce_raw * 100.0

        strength = weighted_score([
            (med_acc_pct, get_weight(growth_weights, "Accumulated Equity Growth", "Accumulated Profit Growth")),
            (median_roe, get_weight(growth_weights, "ROE")),
            (med_roce_pct, get_weight(growth_weights, "ROCE")),
        ])
        volatility = weighted_score([
            (std_acc_pct, get_weight(stddev_weights, "Accumulated Equity Growth", "Accumulated Profit Growth")),
            (std_roe, get_weight(stddev_weights, "ROE")),
            (std_roce_pct, get_weight(stddev_weights, "ROCE")),
        ])
        if strength is not None and volatility is not None and med_interest is not None:
            bs_scaled = strength / (1.0 + volatility)
            bs_debt = bs_scaled / (1.0 + med_interest / 100.0)

    # P&L component.  Combined Dashboard uses the revenue series to establish
    # the most recent year for every P&L submetric.
    if not series["revenue"].empty:
        yr_start, yr_end = int(series["revenue"]["year"].max()), 2020
        med_rev_raw, std_rev_raw = compute_growth_stats(
            series["revenue"], yr_start, yr_end, True, "revenue", True
        )
        med_pt, std_pt = compute_growth_stats(series["pretax"], yr_start, yr_end, True, "pretax_income", True) if not series["pretax"].empty else (None, None)
        med_ni, std_ni = compute_growth_stats(series["net_income"], yr_start, yr_end, True, "net_income", True) if not series["net_income"].empty else (None, None)
        med_nopat, std_nopat = compute_growth_stats(series["nopat"], yr_start, yr_end, True, "nopat", True) if not series["nopat"].empty else (None, None)

        if not series["op_margin"].empty:
            med_om_raw, std_om_raw, margin_is_fraction = compute_margin_stats(series["op_margin"], yr_start, yr_end, True)
            med_om_growth_raw, std_om_growth_raw = compute_margin_growth_stats(series["op_margin"], yr_start, yr_end, True)
            median_om = None if med_om_raw is None else med_om_raw * (100.0 if margin_is_fraction else 1.0)
            std_om = None if std_om_raw is None else std_om_raw * (100.0 if margin_is_fraction else 1.0)
            median_om_growth = None if med_om_growth_raw is None else med_om_growth_raw * 100.0
            std_om_growth = None if std_om_growth_raw is None else std_om_growth_raw * 100.0

        median_rev = None if med_rev_raw is None else med_rev_raw * 100.0
        std_rev = None if std_rev_raw is None else std_rev_raw * 100.0
        med_pt_pct = None if med_pt is None else med_pt * 100.0
        std_pt_pct = None if std_pt is None else std_pt * 100.0
        med_ni_pct = None if med_ni is None else med_ni * 100.0
        std_ni_pct = None if std_ni is None else std_ni * 100.0
        med_nopat_pct = None if med_nopat is None else med_nopat * 100.0
        std_nopat_pct = None if std_nopat is None else std_nopat * 100.0

        growth = weighted_score([
            (median_rev, get_weight(growth_weights, "Revenue Growth")),
            (med_pt_pct, get_weight(growth_weights, "Pretax Income Growth", "Profit Before Tax Growth")),
            (med_ni_pct, get_weight(growth_weights, "Net Income Growth", "Net Income  Growth")),
            (med_nopat_pct, get_weight(growth_weights, "NOPAT Growth")),
            (median_om, get_weight(growth_weights, "Operating Margin")),
            (median_om_growth, get_weight(growth_weights, "YoY Operating Margin Growth")),
        ])
        volatility = weighted_score([
            (std_rev, get_weight(stddev_weights, "Revenue Growth")),
            (std_pt_pct, get_weight(stddev_weights, "Pretax Income Growth", "Profit Before Tax Growth")),
            (std_ni_pct, get_weight(stddev_weights, "Net Income Growth", "Net Income  Growth")),
            (std_nopat_pct, get_weight(stddev_weights, "NOPAT Growth")),
            (std_om, get_weight(stddev_weights, "Operating Margin")),
            (std_om_growth, get_weight(stddev_weights, "YoY Operating Margin Growth")),
        ])
        if growth is not None and volatility is not None:
            pl_scaled = growth / (1.0 + volatility)

    # FCFF/spread component.  Its range starts at the most recent year in the
    # union of the FCFF and ROIC-WACC spread series.
    fs_years: list[int] = []
    for key in ("fcff", "spread"):
        if not series[key].empty:
            fs_years.extend(pd.to_numeric(series[key]["year"], errors="coerce").dropna().astype(int).tolist())
    if fs_years:
        yr_start, yr_end = max(fs_years), 2020
        med_fcff_raw, std_fcff_raw = compute_growth_stats(
            series["fcff"], yr_start, yr_end, True, "fcff", True
        ) if not series["fcff"].empty else (None, None)
        median_fcff = None if med_fcff_raw is None else med_fcff_raw * 100.0
        std_fcff = None if std_fcff_raw is None else std_fcff_raw * 100.0
        median_spread, std_spread = level_stats(series["spread"], yr_start, yr_end, "spread_pct")

        growth = weighted_score([
            (median_fcff, get_weight(growth_weights, "FCFF Growth", "FCFE Growth")),
            (median_spread, get_weight(growth_weights, "Spread")),
        ])
        volatility = weighted_score([
            (std_fcff, get_weight(stddev_weights, "FCFF Growth", "FCFE Growth")),
            (std_spread, get_weight(stddev_weights, "Spread")),
        ])
        if growth is not None and volatility is not None:
            fs_scaled = growth / (1.0 + volatility)

    debt_adjusted_total = (
        bs_debt + pl_scaled + fs_scaled
        if bs_debt is not None and pl_scaled is not None and fs_scaled is not None
        else None
    )

    return {
        "Debt-Adj. Total Score": finite_or_none(debt_adjusted_total),
        "Median ROIC-WACC": finite_or_none(median_spread),
        "ROIC-WACC Std Dev": finite_or_none(std_spread),
        "Median Op. Margin": finite_or_none(median_om),
        "Op. Margin Std Dev": finite_or_none(std_om),
        "Median Op. Margin Growth": finite_or_none(median_om_growth),
        "Op. Margin Growth Std Dev": finite_or_none(std_om_growth),
        "Median Revenue Growth": finite_or_none(median_rev),
        "Median ROE": finite_or_none(median_roe),
        "ROE Std Dev": finite_or_none(std_roe),
        "Median FCFF Change": finite_or_none(median_fcff),
    }


def apply_sheet_style(worksheet, table_name: str, widths: dict[str, float]) -> None:
    header_fill = PatternFill("solid", fgColor="17365D")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    worksheet.row_dimensions[1].height = 34
    worksheet.freeze_panes = "C2" if worksheet.title == "Company Metrics" else "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    for column, width in widths.items():
        worksheet.column_dimensions[column].width = width
    table = Table(displayName=table_name, ref=worksheet.dimensions)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    worksheet.add_table(table)


def write_workbook(
    output_path: Path,
    source_rows: list[dict[str, Any]],
    result_rows: list[dict[str, Any]],
) -> None:
    workbook = Workbook()
    sheet1 = workbook.active
    sheet1.title = "Company Metrics"
    sheet1.append(SHEET1_COLUMNS)
    for result in result_rows:
        sheet1.append([result.get(column) for column in SHEET1_COLUMNS])

    for row in sheet1.iter_rows(min_row=2, min_col=3, max_col=len(SHEET1_COLUMNS)):
        for cell in row:
            cell.number_format = "0.00"
    for row_index, source in enumerate(source_rows, start=2):
        if source.get("company_id") is None:
            sheet1.cell(row=row_index, column=1).comment = Comment(
                "No matching company record was found in the TaRaSha Research database; research fields are blank.",
                "Codex",
            )

    unmatched = [row["ticker"] for row in source_rows if row.get("company_id") is None]
    sheet1["A1"].comment = Comment(
        "Source: TaRaSha Research, Equity Research > Value Creation Stability Score > Combined Dashboard. "
        "Settings: Recent - 2020; sample standard deviation (ddof=1). Percentage-based metrics are "
        "stored as percentage points. "
        + (f"No database match: {', '.join(unmatched)}." if unmatched else "All source companies matched."),
        "Codex",
    )
    apply_sheet_style(
        sheet1,
        "CompanyMetricsTable",
        {"A": 42, "B": 12, "C": 22, "D": 20, "E": 21, "F": 20, "G": 20, "H": 25, "I": 27, "J": 23, "K": 18, "L": 18, "M": 22},
    )

    sheet2 = workbook.create_sheet("Top Quartile Values")
    sheet2.append(["Metric", "Top Quartile Value (75th Percentile)", "Companies with Data"])
    results_df = pd.DataFrame(result_rows)
    for metric in METRIC_COLUMNS:
        values = pd.to_numeric(results_df[metric], errors="coerce").replace([np.inf, -np.inf], np.nan).dropna()
        quartile = float(values.quantile(0.75, interpolation="linear")) if not values.empty else None
        sheet2.append([metric, quartile, int(values.count())])
    for cell in sheet2["B"][1:]:
        cell.number_format = "0.00"
    sheet2["A1"].comment = Comment(
        "Top quartile is the inclusive 75th percentile (equivalent to Excel QUARTILE.INC(array,3)) "
        "calculated from the nonblank values in Company Metrics.",
        "Codex",
    )
    apply_sheet_style(sheet2, "TopQuartileValuesTable", {"A": 34, "B": 38, "C": 22})

    workbook.properties.title = "US-Listed AI Token Cost Beneficiaries - TaRaSha Combined Dashboard Metrics"
    workbook.properties.subject = "Combined Dashboard metrics and 75th-percentile top-quartile cutoffs"
    workbook.properties.creator = "TaRaSha Research"
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("source", type=Path)
    parser.add_argument("output", type=Path)
    args = parser.parse_args()

    source_rows = load_source_companies(args.source)
    session = SessionLocal()
    conn = DbCompat(session)
    try:
        companies = read_df("SELECT id, name, ticker FROM companies ORDER BY id", conn)
        match_company_ids(source_rows, companies)
        company_ids = sorted({int(row["company_id"]) for row in source_rows if row.get("company_id") is not None})

        growth_df = read_df("SELECT factor, weight FROM growth_weight_factors", conn)
        stddev_df = read_df("SELECT factor, weight FROM stddev_weight_factors", conn)
        growth_weights = {str(row["factor"]): float(row["weight"]) for _, row in growth_df.dropna(subset=["weight"]).iterrows()}
        stddev_weights = {str(row["factor"]): float(row["weight"]) for _, row in stddev_df.dropna(subset=["weight"]).iterrows()}
        frames = load_annual_tables(conn, company_ids)

        result_rows: list[dict[str, Any]] = []
        for source in source_rows:
            result: dict[str, Any] = {"Company": source["company"], "Ticker": source["ticker"]}
            if source.get("company_id") is not None:
                result.update(compute_company_metrics(int(source["company_id"]), frames, growth_weights, stddev_weights))
            else:
                result.update({metric: None for metric in METRIC_COLUMNS})
            result_rows.append(result)
    finally:
        session.rollback()
        session.close()

    write_workbook(args.output, source_rows, result_rows)
    unmatched = [row["ticker"] for row in source_rows if row.get("company_id") is None]
    coverage = {
        metric: sum(row.get(metric) is not None for row in result_rows)
        for metric in METRIC_COLUMNS
    }
    print(f"Wrote {args.output}")
    print(f"Source rows: {len(source_rows)}; matched: {len(source_rows) - len(unmatched)}; unmatched: {len(unmatched)}")
    if unmatched:
        print("Unmatched tickers: " + ", ".join(unmatched))
    print("Coverage: " + "; ".join(f"{metric}={count}" for metric, count in coverage.items()))


if __name__ == "__main__":
    main()
